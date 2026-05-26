import pandas as pd
import numpy as np
import os
import glob
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
import warnings
import sys

# ====================== 全局初始化 ======================
warnings.filterwarnings("ignore")
# 切换工作目录为脚本所在目录
os.chdir(os.path.dirname(os.path.abspath(sys.argv[0])))
# 中文显示配置
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

# ====================== 核心配置区（集中管理，精简分区） ======================
CONFIG = {
    # 基础文件配置
    "OUTPUT_EXCEL": "产品检测结果总表.xlsx",
    "ENCODING": "gbk",
    "COL": [1, 4, 10],               # 电流、流量、压差 列索引
    "REF_FLOW": [5.0, 10.0, 20.0, 40.0, 60.0],
    "PLOT_FLOW": 20.0,
    "STD_LIMIT": 10.0,
    "TOLERANCE": 1.00,              # 流量匹配误差阈值
    "FILTER_ON": True,
    "FILTER_WINDOW": 201,

    # 关键参数
    "KEY_FLOW": 20.0,
    "HIGH_FLOW": 40.0,
    "KEY_CURRENTS": [0, 300, 600, 900],
    "CONSISTENCY_LIMIT_PCT": 5.0,

    # 图表通用样式
    "X_AXIS_EXTEND": 8,
    "HIDE_TOP_RIGHT_BORDER": True,
    "PERF_LINE_WIDTH": 2.0,
    "PERF_TITLE_SUFFIX": " 实测",
    "PERF_TITLE_SIZE": 36,
    "PERF_TITLE_BOLD": False,
    "PERF_TITLE_OFFSET_X": 0.5,
    "PERF_TITLE_OFFSET_Y": 1.1,
    "PERF_AXIS_LABEL_SIZE": 20,
    "PERF_AXIS_LABEL_BOLD": False,
    "PERF_TICK_LABEL_SIZE": 20,
    "PERF_TICK_WIDTH": 1.2,
    "PERF_TEXT_OFFSET_X": 1.5,
    "PERF_TEXT_OFFSET_Y": 0.0,
    "PERF_CURR_TEXT_SIZE": 20,
    "PERF_CURR_TEXT_BOLD": False,

    # 一致性图表样式
    "CON_TITLE": "一致性",
    "CON_TITLE_SIZE": 24,
    "CON_TITLE_BOLD": False,
    "CON_TITLE_OFFSET_X": 0.5,
    "CON_TITLE_OFFSET_Y": 1.1,
    "CON_AXIS_LABEL_SIZE": 20,
    "CON_AXIS_LABEL_BOLD": False,
    "CON_TICK_LABEL_SIZE": 16,
    "CON_LINE_WIDTH": 2.0,
    "CON_TICK_WIDTH": 1.2,
    "LABEL_POS_0mA": -16,
    "LABEL_POS_OTHER": 9,
    "CON_LABEL_SIZE": 15,
    "CON_CURR_TEXT_SIZE": 20,

    # 功能总开关
    "RUN_PERFORMANCE_PLOTS": True,
    "RUN_PQ_COMPARE_PLOT": True,
    "RUN_HYSTERESIS_PLOT": True,
    "RUN_CONSISTENCY_PLOT": True,
    "RUN_AVG_PRESSURE_PLOT": True,
    "RUN_EXCEL_CHARTS": True,
    "RUN_EXCEL_REPORT": True,
    "RUN_SINGLE_GOBACK_PQ": True,
    "RUN_ALL_GOBACK_SUMMARY": True,
}

# 全局颜色配置（固定常量）
COLOR_MAP = {0: '#70AD47', 1: '#00B0F0', 2: '#FFC000', 3: '#FF0000', 4: '#C00000'}
COLOR_EXCEL = ["70AD47", "00B0F0", "FFC000", "FF0000", "FF0000", "C00000"]
LINE_COLOR = "#00B0F0"

# ====================== 通用工具函数（统一封装，无冗余） ======================
def create_output_dir(base_name="检测结果输出"):
    """自动创建输出文件夹，重复则编号"""
    out_dir = base_name
    if os.path.exists(out_dir):
        cnt = 1
        while os.path.exists(f"{out_dir}_{cnt}"):
            cnt += 1
        out_dir = f"{out_dir}_{cnt}"
    os.makedirs(out_dir)
    return out_dir

def safe_save_fig(fig, out_dir, filename, dpi=150):
    """安全保存图片，自动重命名防覆盖"""
    path = os.path.join(out_dir, filename)
    base, ext = os.path.splitext(path)
    counter = 1
    while os.path.exists(path):
        path = f"{base}({counter}){ext}"
        counter += 1
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)

def closest_idx(arr, target):
    """查找数组中最接近目标值的索引"""
    return np.argmin(np.abs(arr - target))

def clean_data(df):
    """统一数据清洗：空值、无穷值、类型转换"""
    df = df.replace([False, None, np.inf, -np.inf], np.nan)
    df = df.dropna(subset=["I", "Q", "P"])
    df["I"] = pd.to_numeric(df["I"], errors="coerce")
    df["Q"] = pd.to_numeric(df["Q"], errors="coerce")
    df["P"] = pd.to_numeric(df["P"], errors="coerce")
    return df

def smooth_press(df):
    """按电流分组，压差滑动平均滤波"""
    if not CONFIG["FILTER_ON"]:
        return df
    df_new = df.copy()
    for curr, group in df.groupby("I"):
        df_new.loc[group.index, "P"] = group["P"].rolling(
            CONFIG["FILTER_WINDOW"], center=True, min_periods=1
        ).mean()
    return df_new

def safe_percent(numerator, denominator, default=0):
    """安全计算百分比，避免除零"""
    try:
        return round((numerator / denominator) * 100, 1)
    except ZeroDivisionError:
        return default

# ====================== 加权一致性核心算法 ======================
def calc_weighted_influence(df_valid):
    """计算产品加权偏差、影响占比、修正系数"""
    KEY_CURR = CONFIG["KEY_CURRENTS"]
    KEY_FLOW = CONFIG["KEY_FLOW"]
    HIGH_FLOW = CONFIG["HIGH_FLOW"]

    group_mean = df_valid.groupby(["I", "参考流量值", "数据分区"])["实际压差"].mean().reset_index()
    group_mean.rename(columns={"实际压差": "工况均值"}, inplace=True)
    df_merge = df_valid.merge(group_mean, on=["I", "参考流量值", "数据分区"], how="left")

    influence = []
    for _, row in df_merge.iterrows():
        I, rf, part = row["I"], row["参考流量值"], row["数据分区"]
        file = row["数据源文件"]
        P, mean_ir = row["实际压差"], row["工况均值"]
        dev = P - mean_ir

        # 跳过指定工况
        if I in [0.0, 300.0] and rf in [5.0, 10.0]:
            continue

        # 权重分配
        if I in KEY_CURR:
            weight_map = {0:10.0, 300:8.0, 600:6.0, 900:4.0}
            base_w = weight_map.get(I, 1.0)
            weight = base_w * 2.0 if rf == KEY_FLOW else base_w
        else:
            weight = 1.0

        influence.append({
            "file": file, "I": I, "rf": rf, "P": P,
            "mean_ir": mean_ir, "dev": dev, "weight": weight,
            "weighted_dev": dev * weight
        })

    if not influence:
        return None, None, None, None

    df_inf = pd.DataFrame(influence)
    file_wdev = df_inf.groupby("file")["weighted_dev"].sum()
    file_weight = df_inf.groupby("file")["weight"].sum()
    file_avg_dev = file_wdev / file_weight
    file_score = file_avg_dev.abs().sort_values(ascending=False)

    # 影响占比
    total_abs = file_score.sum()
    impact_rate = {f: round(s / total_abs * 100, 1) for f, s in file_score.items()}

    # 修正系数
    global_mean = df_inf["mean_ir"].mean()
    scale_suggest = {}
    for f in file_avg_dev.index:
        f_mean = df_inf[df_inf["file"] == f]["P"].mean()
        adj_mean = f_mean + file_avg_dev[f]
        scale = round(global_mean / adj_mean, 3) if adj_mean != 0 else 1.0
        scale_suggest[f] = scale

    worst_file = file_score.index[0] if len(file_score) > 0 else None
    return file_score, impact_rate, worst_file, scale_suggest

def get_0mA60L_data(df_valid, file_list):
    """提取0mA 60L/min基准压差"""
    res = {}
    target_I, target_rf = 0.0, 60.0
    for f in file_list:
        sub = df_valid[
            (df_valid["数据源文件"] == f) & (df_valid["I"] == target_I) &
            (df_valid["参考流量值"] == target_rf) & (df_valid["数据分区"] == "前50%")
        ]
        res[f] = round(sub["实际压差"].iloc[0], 2) if not sub.empty else 0.0
    return res

# ====================== 主程序流程（线性结构，清晰直观） ======================
if __name__ == "__main__":
    print("=" * 70)
    print("🚀 PQ 产品一致性检测分析程序 启动")
    print("=" * 70)

    # 1. 创建输出目录
    OUTPUT = create_output_dir("检测结果输出")
    print(f"\n📂 结果输出目录：{OUTPUT}")

    # 2. 扫描并读取CSV文件
    print("\n【1/12】扫描CSV文件...")
    csv_files = glob.glob("*.csv")
    print(f"✅ 共找到 {len(csv_files)} 个CSV文件")

    raw_data = []
    pq_compare_data = []

    for file in csv_files:
        try:
            # 读取指定列并重命名
            df = pd.read_csv(file, usecols=CONFIG["COL"], encoding=CONFIG["ENCODING"])
            df.columns = ["I", "Q", "P"]
            df = clean_data(df)
            if df.empty:
                continue

            # 数据滤波
            df = smooth_press(df)
            file_short = os.path.splitext(file)[0]

            # 提取PQ对比数据（仅去程）
            for curr, group in df.groupby("I"):
                curr = round(curr, 2)
                mid = len(group) // 2
                go_part = group.iloc[:mid]
                go_part = clean_data(go_part)
                if go_part.empty:
                    continue

                q_arr = go_part["Q"].abs().values
                p_arr = go_part["P"].abs().values
                q_points, p_points = [], []
                for ref in CONFIG["REF_FLOW"]:
                    idx = closest_idx(q_arr, ref)
                    q_points.append(round(abs(q_arr[idx]), 4))
                    p_points.append(round(abs(p_arr[idx]), 4))

                pq_compare_data.append({
                    "file": file_short, "current": curr,
                    "flow": q_points, "press": p_points
                })

            # 拆分去程/回程，匹配标准流量点
            for curr, group in df.groupby("I"):
                curr = round(curr, 2)
                mid = len(group) // 2
                for part_name, part_data in [("前50%", group.iloc[:mid]), ("后50%", group.iloc[mid:])]:
                    part_data = clean_data(part_data)
                    if part_data.empty:
                        continue
                    q_arr = part_data["Q"].abs().values
                    p_arr = part_data["P"].abs().values

                    for ref_flow in CONFIG["REF_FLOW"]:
                        idx = closest_idx(q_arr, ref_flow)
                        real_q = round(abs(q_arr[idx]), 4)
                        real_p = round(abs(p_arr[idx]), 4)
                        diff_q = round(abs(real_q - ref_flow), 4)
                        check_flag = "TRUE" if diff_q <= CONFIG["TOLERANCE"] else "FALSE"
                        raw_data.append([
                            file, curr, part_name, ref_flow, real_p, real_q, diff_q, check_flag
                        ])
        except Exception as e:
            print(f"⚠️ 文件 {file} 读取失败：{str(e)}")
            continue

    # 3. 构建基础数据表 & 筛选全文件有效点位
    print("\n【2/12】数据预处理与筛选...")
    df_raw = pd.DataFrame(raw_data, columns=[
        "数据源文件", "I", "数据分区", "参考流量值", "实际压差", "实际流量", "流量差值", "校验结果"
    ])

    # 仅保留所有文件都校验通过的点位
    group_check = df_raw.groupby(["I", "参考流量值", "数据分区"])["校验结果"].apply(lambda x: (x == "TRUE").all())
    valid_groups = group_check[group_check].index.tolist()
    df_valid = df_raw[df_raw.set_index(["I", "参考流量值", "数据分区"]).index.isin(valid_groups)].copy()
    print("✅ 数据筛选完成")

    # 4. 加权一致性分析
    print("\n【3/12】加权一致性分析...")
    file_score, impact_rate, worst_file, scale_suggest = calc_weighted_influence(df_valid)
    all_files = df_valid["数据源文件"].unique()
    orig_0mA60L = get_0mA60L_data(df_valid, all_files)
    scale_0mA60L = {f: (orig_0mA60L[f], round(orig_0mA60L[f] * scale_suggest.get(f, 1.0), 2)) for f in all_files}
    print("✅ 一致性分析完成")

    # 5. 计算迟滞数据
    print("\n【4/12】计算迟滞数据...")
    hys_data = []
    for (f, curr, rf), group in df_valid.groupby(["数据源文件", "I", "参考流量值"]):
        go_p = group[group["数据分区"] == "前50%"]["实际压差"]
        back_p = group[group["数据分区"] == "后50%"]["实际压差"]
        if len(go_p) == 1 and len(back_p) == 1:
            hys = round(abs(go_p.iloc[0] - back_p.iloc[0]), 2)
            hys_data.append([f, curr, rf, hys])
    df_hys = pd.DataFrame(hys_data, columns=["数据源文件", "I", "参考流量值", "迟滞"])
    print("✅ 迟滞计算完成")

    # 6. 计算一致性统计数据
    print("\n【5/12】计算一致性统计数据...")
    con_data = []
    for (curr, rf, part), group in df_valid.groupby(["I", "参考流量值", "数据分区"]):
        p_vals = group["实际压差"].dropna()
        if len(p_vals) < 1:
            continue
        max_p = p_vals.max()
        min_p = p_vals.min()
        avg_p = (max_p + min_p) / 2
        diff_half = (max_p - min_p) / 2
        max_file = group.loc[p_vals.idxmax(), "数据源文件"]
        min_file = group.loc[p_vals.idxmin(), "数据源文件"]
        pct = safe_percent(diff_half, avg_p)
        con_data.append([
            round(curr, 2), round(rf, 1), part, max_p, max_file,
            min_p, min_file, avg_p, diff_half, pct
        ])
    df_con = pd.DataFrame(con_data, columns=[
        "I", "参考流量值", "数据分区", "最大值", "最大值文件",
        "最小值", "最小值文件", "平均值", "半差值", "一致性%"
    ])
    print("✅ 一致性统计完成")

    # ====================== 批量生成图表 ======================
    print("\n【6/12】开始生成各类图表...")

    # 6.1 单文件性能图
    if CONFIG["RUN_PERFORMANCE_PLOTS"]:
        perf_dir = os.path.join(OUTPUT, "性能图表")
        os.makedirs(perf_dir, exist_ok=True)
        file_list = df_valid["数据源文件"].unique()
        for idx, f in enumerate(file_list, 1):
            sub_df = df_valid[(df_valid["数据源文件"] == f) & (df_valid["数据分区"] == "前50%")]
            if sub_df.empty:
                continue
            fig, ax = plt.subplots(figsize=(12, 6), dpi=100)
            curr_list = sorted(sub_df["I"].unique())
            for curr in curr_list:
                curr_df = sub_df[sub_df["I"] == curr].sort_values("参考流量值")
                ax.plot(curr_df["参考流量值"], curr_df["实际压差"], marker="o", ms=4,
                        color=LINE_COLOR, lw=CONFIG["PERF_LINE_WIDTH"])
                # 电流标注
                x_last = curr_df["参考流量值"].iloc[-1]
                y_last = curr_df["实际压差"].iloc[-1]
                ax.text(x_last + CONFIG["PERF_TEXT_OFFSET_X"], y_last + CONFIG["PERF_TEXT_OFFSET_Y"],
                        f"{curr}mA", fontsize=CONFIG["PERF_CURR_TEXT_SIZE"],
                        weight="bold" if CONFIG["PERF_CURR_TEXT_BOLD"] else "normal")
            # 样式设置
            title = f"{os.path.splitext(f)[0]}{CONFIG['PERF_TITLE_SUFFIX']}"
            ax.set_title(title, fontsize=CONFIG["PERF_TITLE_SIZE"],
                         weight="bold" if CONFIG["PERF_TITLE_BOLD"] else "normal",
                         x=CONFIG["PERF_TITLE_OFFSET_X"], y=CONFIG["PERF_TITLE_OFFSET_Y"])
            ax.set_xlabel("流量 L/min", fontsize=CONFIG["PERF_AXIS_LABEL_SIZE"],
                          weight="bold" if CONFIG["PERF_AXIS_LABEL_BOLD"] else "normal")
            ax.set_ylabel("压差 bar", fontsize=CONFIG["PERF_AXIS_LABEL_SIZE"],
                          weight="bold" if CONFIG["PERF_AXIS_LABEL_BOLD"] else "normal")
            ax.tick_params(axis="both", labelsize=CONFIG["PERF_TICK_LABEL_SIZE"], width=CONFIG["PERF_TICK_WIDTH"])
            ax.set_xlim(right=70)
            if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
                ax.spines["top"].set_visible(False)
                ax.spines["right"].set_visible(False)
            plt.tight_layout()
            safe_save_fig(fig, perf_dir, f"性能图表-{idx}.png")
        print("✅ 单文件性能图生成完成")

    # 6.2 PQ总对比图
    if CONFIG["RUN_PQ_COMPARE_PLOT"] and pq_compare_data:
        fig, ax = plt.subplots(figsize=(18, 10), dpi=300)
        fig.set_facecolor("white")
        unique_files = sorted({d["file"] for d in pq_compare_data})
        unique_currs = sorted({d["current"] for d in pq_compare_data})
        import matplotlib.cm as cm
        colors = cm.nipy_spectral(np.linspace(0, 1, len(unique_files)))
        file_color_map = dict(zip(unique_files, colors))

        plotted = set()
        for f in unique_files:
            f_data = [d for d in pq_compare_data if d["file"] == f]
            for curr in unique_currs:
                curr_data = next((d for d in f_data if d["current"] == curr), None)
                if not curr_data:
                    continue
                lab = f if f not in plotted else ""
                if lab:
                    plotted.add(f)
                ax.plot(curr_data["flow"], curr_data["press"], marker="o", ms=3,
                        linewidth=0.6, color=file_color_map[f], label=lab, alpha=0.8)
        # 标注电流
        if unique_currs:
            first_data = [d for d in pq_compare_data if d["file"] == unique_files[0]]
            for curr in unique_currs:
                cd = next((d for d in first_data if d["current"] == curr), None)
                if cd:
                    x, y = cd["flow"][-1], cd["press"][-1]
                    ax.text(x + CONFIG["PERF_TEXT_OFFSET_X"], y, f"{curr}mA",
                            fontsize=CONFIG["CON_CURR_TEXT_SIZE"])
        # 样式
        ax.set_title("所有产品滤波后前50% PQ曲线对比", fontsize=16, weight="bold")
        ax.set_xlabel("流量 L/min", fontsize=14)
        ax.set_ylabel("压差 bar", fontsize=14)
        ax.grid(alpha=0.3)
        ax.set_xlim(0, max(CONFIG["REF_FLOW"]) + 10)
        ax.tick_params(labelsize=12)
        if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
        ncol = min(3, max(1, len(unique_files) // 12))
        ax.legend(loc="center left", bbox_to_anchor=(1.03, 0.5), ncol=ncol, fontsize=10)
        plt.tight_layout(rect=[0, 0, 0.87, 1])
        safe_save_fig(fig, OUTPUT, "所有产品PQ曲线对比图.png", dpi=300)
        print("✅ PQ总对比图生成完成")

    # 6.3 迟滞PNG图
    if CONFIG["RUN_HYSTERESIS_PLOT"]:
        plot_hys = df_hys[df_hys["参考流量值"] == CONFIG["PLOT_FLOW"]].sort_values(["数据源文件", "I"])
        fig, ax = plt.subplots(figsize=(15, 6), dpi=100)
        for f in plot_hys["数据源文件"].unique():
            sub = plot_hys[plot_hys["数据源文件"] == f]
            exceed = sum(sub["迟滞"] > CONFIG["STD_LIMIT"])
            color = COLOR_MAP[min(exceed, 4)]
            ax.plot(sub["I"], sub["迟滞"], marker="o", ms=3, lw=1, color=color, label=f)
        # 标准线
        ax.axhline(CONFIG["STD_LIMIT"], color="r", linestyle="--", lw=1, label=f"标准线 {CONFIG['STD_LIMIT']}")
        ax.set_title("P-Q滞环@20L/min")
        ax.set_xlabel("电流值")
        ax.set_ylabel("迟滞/bar")
        ax.grid(axis="y")
        ncol = min(3, len(plot_hys["数据源文件"].unique()) // 15 + 1)
        ax.legend(loc="center left", bbox_to_anchor=(1.02, 0.5), ncol=ncol, fontsize=8)
        if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
        plt.tight_layout()
        safe_save_fig(fig, OUTPUT, "迟滞图.png")
        print("✅ 迟滞图生成完成")

    # 6.4 一致性图 & 压差平均值图
    if CONFIG["RUN_CONSISTENCY_PLOT"]:
        con_plot = df_con[df_con["数据分区"] == "前50%"]
        curr_list = sorted(con_plot["I"].unique())
        min_two_curr = curr_list[:2]

        fig, ax = plt.subplots(figsize=(15, 6), dpi=100)
        for curr in curr_list:
            sub = con_plot[con_plot["I"] == curr].sort_values("参考流量值")
            ax.plot(sub["参考流量值"], sub["平均值"], marker="o", ms=4, color=LINE_COLOR, lw=CONFIG["CON_LINE_WIDTH"])
            # 电流标注
            x_last = sub["参考流量值"].iloc[-1]
            y_last = sub["平均值"].iloc[-1]
            ax.text(x_last + CONFIG["PERF_TEXT_OFFSET_X"], y_last, f"{curr}mA",
                    fontsize=CONFIG["CON_CURR_TEXT_SIZE"],
                    weight="bold" if CONFIG["CON_TITLE_BOLD"] else "normal")
            # 误差标注
            for x, y, dif, pct in zip(sub["参考流量值"], sub["平均值"], sub["半差值"], sub["一致性%"]):
                if curr in min_two_curr and x in [5.0, 10.0]:
                    lab = f"±{dif:.2f}bar"
                else:
                    lab = f"±{pct:.2f}%"
                offset = CONFIG["LABEL_POS_0mA"] if curr == 0 else CONFIG["LABEL_POS_OTHER"]
                ax.annotate(lab, (x, y), xytext=(0, offset), textcoords="offset points",
                            ha="center", fontsize=CONFIG["CON_LABEL_SIZE"])
        # 样式
        ax.set_title(CONFIG["CON_TITLE"], fontsize=CONFIG["CON_TITLE_SIZE"],
                     weight="bold" if CONFIG["CON_TITLE_BOLD"] else "normal",
                     x=CONFIG["CON_TITLE_OFFSET_X"], y=CONFIG["CON_TITLE_OFFSET_Y"])
        ax.set_xlabel("流量/Lpm", fontsize=CONFIG["CON_AXIS_LABEL_SIZE"],
                      weight="bold" if CONFIG["CON_AXIS_LABEL_BOLD"] else "normal")
        ax.set_ylabel("压差 bar", fontsize=CONFIG["CON_AXIS_LABEL_SIZE"],
                      weight="bold" if CONFIG["CON_AXIS_LABEL_BOLD"] else "normal")
        ax.tick_params(axis="both", labelsize=CONFIG["CON_TICK_LABEL_SIZE"], width=CONFIG["CON_TICK_WIDTH"])
        ax.set_xlim(right=ax.get_xlim()[1] + CONFIG["X_AXIS_EXTEND"])
        if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
        plt.tight_layout()
        safe_save_fig(fig, OUTPUT, "一致性图.png")
        print("✅ 一致性图生成完成")

    if CONFIG["RUN_AVG_PRESSURE_PLOT"]:
        con_avg = df_con[df_con["数据分区"] == "前50%"]
        curr_avg_list = sorted(con_avg["I"].unique())
        fig, ax = plt.subplots(figsize=(15, 6), dpi=100)
        for curr in curr_avg_list:
            sub = con_avg[con_avg["I"] == curr].sort_values("参考流量值")
            ax.plot(sub["参考流量值"], sub["平均值"], marker="o", ms=4, color=LINE_COLOR, lw=CONFIG["CON_LINE_WIDTH"])
            x_last = sub["参考流量值"].iloc[-1]
            y_last = sub["平均值"].iloc[-1]
            ax.text(x_last + CONFIG["PERF_TEXT_OFFSET_X"], y_last, f"{curr}mA",
                    fontsize=CONFIG["CON_CURR_TEXT_SIZE"])
            # 数值标注
            for x, y, avg, dif in zip(sub["参考流量值"], sub["平均值"], sub["平均值"], sub["半差值"]):
                lab = f"{avg:.2f}±{dif:.2f} bar"
                offset = CONFIG["LABEL_POS_0mA"] if curr == 0 else CONFIG["LABEL_POS_OTHER"]
                ax.annotate(lab, (x, y), xytext=(0, offset), textcoords="offset points",
                            ha="center", fontsize=CONFIG["CON_LABEL_SIZE"])
        ax.set_title("压差平均值", fontsize=CONFIG["CON_TITLE_SIZE"],
                     x=CONFIG["CON_TITLE_OFFSET_X"], y=CONFIG["CON_TITLE_OFFSET_Y"])
        ax.set_xlabel("流量/Lpm", fontsize=CONFIG["CON_AXIS_LABEL_SIZE"])
        ax.set_ylabel("压差 bar", fontsize=CONFIG["CON_AXIS_LABEL_SIZE"])
        ax.tick_params(axis="both", labelsize=CONFIG["CON_TICK_LABEL_SIZE"], width=CONFIG["CON_TICK_WIDTH"])
        ax.set_xlim(right=ax.get_xlim()[1] + CONFIG["X_AXIS_EXTEND"])
        if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
        plt.tight_layout()
        safe_save_fig(fig, OUTPUT, "压差平均值图.png")
        print("✅ 压差平均值图生成完成")

    # 6.5 去程/回程 单文件&总对比图
    all_go_data, all_back_data, all_both_data = [], [], []
    if CONFIG["RUN_SINGLE_GOBACK_PQ"]:
        dir_go = os.path.join(OUTPUT, "01_去程PQ")
        dir_back = os.path.join(OUTPUT, "02_回程PQ")
        dir_both = os.path.join(OUTPUT, "03_去程+回程PQ")
        for d in [dir_go, dir_back, dir_both]:
            os.makedirs(d, exist_ok=True)

        for idx, file in enumerate(csv_files, 1):
            try:
                df = pd.read_csv(file, usecols=CONFIG["COL"], encoding=CONFIG["ENCODING"])
                df.columns = ["I", "Q", "P"]
                df = clean_data(df)
                df = smooth_press(df)
                file_short = os.path.splitext(file)[0]
                curr_list = sorted(df["I"].unique())
                go_list, back_list = [], []

                # 去程图
                fig1, ax1 = plt.subplots(figsize=(12, 6), dpi=100)
                for curr in curr_list:
                    sub = df[df["I"] == curr]
                    go = sub.iloc[:len(sub)//2].sort_values("Q")
                    x = go["Q"].abs()
                    y = go["P"].abs()
                    if len(x) >= 2:
                        ax1.plot(x, y, lw=CONFIG["PERF_LINE_WIDTH"], color=LINE_COLOR)
                        ax1.text(x.iloc[-1] + CONFIG["PERF_TEXT_OFFSET_X"], y.iloc[-1], f"{round(curr)}mA",
                                fontsize=CONFIG["PERF_CURR_TEXT_SIZE"])
                        go_list.append((x, y))
                ax1.set_title(f"{file_short} 去程PQ", fontsize=CONFIG["PERF_TITLE_SIZE"],
                             x=CONFIG["PERF_TITLE_OFFSET_X"], y=CONFIG["PERF_TITLE_OFFSET_Y"])
                ax1.set_xlabel("流量 L/min", fontsize=CONFIG["PERF_AXIS_LABEL_SIZE"])
                ax1.set_ylabel("压差 bar", fontsize=CONFIG["PERF_AXIS_LABEL_SIZE"])
                ax1.tick_params(labelsize=CONFIG["PERF_TICK_LABEL_SIZE"])
                ax1.set_xlim(right=70)
                if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
                    ax1.spines["top"].set_visible(False)
                    ax1.spines["right"].set_visible(False)
                plt.tight_layout()
                safe_save_fig(fig1, dir_go, f"去程PQ_{idx}.png")

                # 回程图
                fig2, ax2 = plt.subplots(figsize=(12, 6), dpi=100)
                for curr in curr_list:
                    sub = df[df["I"] == curr]
                    back = sub.iloc[len(sub)//2:].sort_values("Q")
                    x = back["Q"].abs()
                    y = back["P"].abs()
                    if len(x) >= 2:
                        ax2.plot(x, y, lw=CONFIG["PERF_LINE_WIDTH"], color=LINE_COLOR)
                        ax2.text(x.iloc[-1] + CONFIG["PERF_TEXT_OFFSET_X"], y.iloc[-1], f"{round(curr)}mA",
                                fontsize=CONFIG["PERF_CURR_TEXT_SIZE"])
                        back_list.append((x, y))
                ax2.set_title(f"{file_short} 回程PQ", fontsize=CONFIG["PERF_TITLE_SIZE"],
                             x=CONFIG["PERF_TITLE_OFFSET_X"], y=CONFIG["PERF_TITLE_OFFSET_Y"])
                ax2.set_xlabel("流量 L/min", fontsize=CONFIG["PERF_AXIS_LABEL_SIZE"])
                ax2.set_ylabel("压差 bar", fontsize=CONFIG["PERF_AXIS_LABEL_SIZE"])
                ax2.tick_params(labelsize=CONFIG["PERF_TICK_LABEL_SIZE"])
                ax2.set_xlim(right=70)
                if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
                    ax2.spines["top"].set_visible(False)
                    ax2.spines["right"].set_visible(False)
                plt.tight_layout()
                safe_save_fig(fig2, dir_back, f"回程PQ_{idx}.png")

                # 合并图
                fig3, ax3 = plt.subplots(figsize=(12, 6), dpi=100)
                for curr in curr_list:
                    sub = df[df["I"] == curr]
                    go = sub.iloc[:len(sub)//2].sort_values("Q")
                    back = sub.iloc[len(sub)//2:].sort_values("Q")
                    xg, yg = go["Q"].abs(), go["P"].abs()
                    xb, yb = back["Q"].abs(), back["P"].abs()
                    if len(xg)>=2:
                        ax3.plot(xg, yg, lw=CONFIG["PERF_LINE_WIDTH"], color=LINE_COLOR)
                        ax3.text(xg.iloc[-1] + CONFIG["PERF_TEXT_OFFSET_X"], yg.iloc[-1], f"{round(curr)}mA",
                                fontsize=CONFIG["PERF_CURR_TEXT_SIZE"])
                    if len(xb)>=2:
                        ax3.plot(xb, yb, lw=CONFIG["PERF_LINE_WIDTH"], color=LINE_COLOR)
                ax3.set_title(f"{file_short} 去程+回程PQ", fontsize=CONFIG["PERF_TITLE_SIZE"],
                             x=CONFIG["PERF_TITLE_OFFSET_X"], y=CONFIG["PERF_TITLE_OFFSET_Y"])
                ax3.set_xlabel("流量 L/min", fontsize=CONFIG["PERF_AXIS_LABEL_SIZE"])
                ax3.set_ylabel("压差 bar", fontsize=CONFIG["PERF_AXIS_LABEL_SIZE"])
                ax3.tick_params(labelsize=CONFIG["PERF_TICK_LABEL_SIZE"])
                ax3.set_xlim(right=70)
                if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
                    ax3.spines["top"].set_visible(False)
                    ax3.spines["right"].set_visible(False)
                plt.tight_layout()
                safe_save_fig(fig3, dir_both, f"去程回程对比_{idx}.png")

                all_go_data.append(go_list)
                all_back_data.append(back_list)
                all_both_data.append((go_list, back_list))
            except Exception as e:
                print(f"⚠️ {file} 图表生成异常：{e}")
                plt.close("all")
        print("✅ 单文件去程/回程图表生成完成")

    # 全体去程回程总对比
    if CONFIG["RUN_ALL_GOBACK_SUMMARY"] and CONFIG["RUN_SINGLE_GOBACK_PQ"]:
        # 去程总对比
        fig, ax = plt.subplots(figsize=(15, 8), dpi=150)
        for d in all_go_data:
            for x, y in d:
                ax.plot(x, y, color=LINE_COLOR, lw=1.0)
        ax.set_title("所有文件 去程PQ总对比", fontsize=20, pad=20)
        ax.set_xlabel("流量 L/min", fontsize=14)
        ax.set_ylabel("压差 bar", fontsize=14)
        ax.set_xlim(right=70)
        if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
        plt.tight_layout()
        safe_save_fig(fig, os.path.join(OUTPUT, "01_去程PQ"), "所有文件_去程总对比.png", dpi=150)

        # 回程总对比
        fig, ax = plt.subplots(figsize=(15, 8), dpi=150)
        for d in all_back_data:
            for x, y in d:
                ax.plot(x, y, color=LINE_COLOR, lw=1.0)
        ax.set_title("所有文件 回程PQ总对比", fontsize=20, pad=20)
        ax.set_xlabel("流量 L/min", fontsize=14)
        ax.set_ylabel("压差 bar", fontsize=14)
        ax.set_xlim(right=70)
        if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
        plt.tight_layout()
        safe_save_fig(fig, os.path.join(OUTPUT, "02_回程PQ"), "所有文件_回程总对比.png", dpi=150)

        # 合并总对比
        fig, ax = plt.subplots(figsize=(15, 8), dpi=150)
        for go_d, back_d in all_both_data:
            for x, y in go_d:
                ax.plot(x, y, color=LINE_COLOR, lw=1.0)
            for x, y in back_d:
                ax.plot(x, y, color=LINE_COLOR, lw=1.0)
        ax.set_title("所有文件 去程+回程PQ总对比", fontsize=20, pad=20)
        ax.set_xlabel("流量 L/min", fontsize=14)
        ax.set_ylabel("压差 bar", fontsize=14)
        ax.set_xlim(right=70)
        if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
        plt.tight_layout()
        safe_save_fig(fig, os.path.join(OUTPUT, "03_去程+回程PQ"), "所有文件_去程回程总对比.png", dpi=150)
        print("✅ 全体去程回程总对比图生成完成")

    # ====================== 生成Excel报告（含内嵌图表） ======================
    print("\n【7/12】生成Excel检测报告...")
    if CONFIG["RUN_EXCEL_REPORT"]:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # 工作表1：原始数据
        ws1 = wb.create_sheet("原始检测数据")
        ws1.append(list(df_raw.columns))
        for _, row in df_raw.iterrows():
            ws1.append(row.tolist())

        # 工作表2：迟滞数据 + Excel图表
        ws2 = wb.create_sheet("产品迟滞")
        ws2.append(list(df_hys.columns))
        for row in hys_data:
            ws2.append(row)

        plot_hys_excel = df_hys[df_hys["参考流量值"] == CONFIG["PLOT_FLOW"]].sort_values(["数据源文件", "I"])
        ws2.cell(1, 8, "绘图文件名")
        ws2.cell(1, 9, "绘图电流")
        ws2.cell(1, 10, "绘图迟滞")
        for i, (_, r) in enumerate(plot_hys_excel.iterrows(), 2):
            ws2.cell(i, 8, r["数据源文件"])
            ws2.cell(i, 9, r["I"])
            ws2.cell(i, 10, r["迟滞"])
        plot_end_row = len(plot_hys_excel) + 1

        # Excel内嵌迟滞图
        if CONFIG["RUN_EXCEL_CHARTS"]:
            chart_hys = ScatterChart()
            chart_hys.scatterStyle = "lineMarker"
            chart_hys.title = "P-Q滞环@20L/min"
            chart_hys.x_axis.title = "电流值"
            chart_hys.y_axis.title = "迟滞/bar"
            prod_list = list(plot_hys_excel["数据源文件"].unique())
            for prod in prod_list:
                sr = er = None
                for row in range(2, plot_end_row + 1):
                    if ws2.cell(row, 8).value == prod:
                        sr = row if sr is None else row
                        er = row
                if not sr:
                    continue
                x_ref = Reference(ws2, 9, sr, 9, er)
                y_ref = Reference(ws2, 10, sr, 10, er)
                # 统计超标点数
                exceed = 0
                for r in range(sr, er + 1):
                    val = ws2.cell(r, 10).value
                    if val and float(val) > CONFIG["STD_LIMIT"]:
                        exceed += 1
                color = COLOR_EXCEL[min(exceed, 4)]
                series = Series(y_ref, x_ref, title=prod)
                series.marker = Marker(size=3)
                series.graphicalProperties.line = LineProperties(solidFill=color, w=10000)
                chart_hys.series.append(series)
            ws2.add_chart(chart_hys, "L2")

        # 工作表3：一致性数据 + Excel图表
        ws3 = wb.create_sheet("一致性对比")
        ws3.append(list(df_con.columns))
        for row in con_data:
            ws3.append(row)

        # 绘制一致性图表通用函数
        def add_con_chart(sheet, part_name, title, start_col, chart_pos):
            data = df_con[df_con["数据分区"] == part_name].sort_values(["I", "参考流量值"])
            headers = ["电流值", "参考流量值", "最大值", "最大值文件", "最小值",
                       "最小值文件", "平均值", "半差值", "一致性%", "数据标签"]
            for idx, h in enumerate(headers):
                sheet.cell(1, start_col + idx, h)
            last_row = 1
            for i, (_, r) in enumerate(data.iterrows(), 2):
                curr = r["I"]
                rf = r["参考流量值"]
                dif = r["半差值"]
                pct = r["一致性%"]
                lab = f"±{dif:.2f}bar" if (curr in [0, 300] and rf in [5, 10]) else f"±{pct:.2f}%"
                sheet.cell(i, start_col, curr)
                sheet.cell(i, start_col + 1, rf)
                sheet.cell(i, start_col + 2, r["最大值"])
                sheet.cell(i, start_col + 3, r["最大值文件"])
                sheet.cell(i, start_col + 4, r["最小值"])
                sheet.cell(i, start_col + 5, r["最小值文件"])
                sheet.cell(i, start_col + 6, r["平均值"])
                sheet.cell(i, start_col + 7, dif)
                sheet.cell(i, start_col + 8, pct)
                sheet.cell(i, start_col + 9, lab)
                last_row = i
            # 创建图表
            chart = ScatterChart()
            chart.scatterStyle = "lineMarker"
            chart.title = f"一致性 {title}"
            chart.x_axis.title = "流量/Lpm"
            chart.y_axis.title = "压差/bar"
            chart.x_axis.majorGridlines = None
            chart.y_axis.majorGridlines = None
            curr_list = sorted(data["I"].unique())
            for curr in curr_list:
                sr = er = None
                for row in range(2, last_row + 1):
                    if sheet.cell(row, start_col).value == curr:
                        sr = row if sr is None else row
                        er = row
                if not sr:
                    continue
                x_ref = Reference(sheet, start_col + 1, sr, start_col + 1, er)
                y_ref = Reference(sheet, start_col + 6, sr, start_col + 6, er)
                label_ref = Reference(sheet, start_col + 9, sr, start_col + 9, er)
                series = Series(y_ref, x_ref, title=f"{curr}mA")
                series.marker = Marker(size=4)
                series.graphicalProperties.line = LineProperties(solidFill="00B0F0", w=8000)
                series.marker.graphicalProperties.solidFill = "00B0F0"
                series.dLbl = True
                series.labelRef = label_ref
                chart.series.append(series)
            sheet.add_chart(chart, chart_pos)

        if CONFIG["RUN_EXCEL_CHARTS"]:
            add_con_chart(ws3, "前50%", "前50%", 15, "AQ2")
            add_con_chart(ws3, "后50%", "后50%", 28, "AQ20")

        # 工作表4：加权一致性分析报告
        ws4 = wb.create_sheet("一致性影响分析(加权)")
        ws4.append(["文件名", "加权偏差总分", "影响占比(%)", "建议缩放系数P", "0mA 60L/min 修正(原→新)", "调节建议"])
        if file_score is not None:
            for f, score in file_score.items():
                rate = impact_rate.get(f, 0)
                scale = scale_suggest.get(f, 1.0)
                orig_p, new_p = scale_0mA60L.get(f, (0.0, 0.0))
                ws4.append([f, round(score, 2), f"{rate}%", scale, f"{orig_p} → {new_p}", f"全量程 × {scale}"])

        # 保存Excel
        excel_path = os.path.join(OUTPUT, CONFIG["OUTPUT_EXCEL"])
        base, ext = os.path.splitext(excel_path)
        cnt = 1
        while os.path.exists(excel_path):
            excel_path = f"{base}({cnt}){ext}"
            cnt += 1
        wb.save(excel_path)
        print("✅ Excel报告生成完成")

    # ====================== 最终结果输出 ======================
    print("\n" + "=" * 85)
    print("📊 产品一致性检测结果（影响排序）")
    print("=" * 85)
    if file_score is not None and len(file_score) > 0:
        print(f"{'文件名':<25} {'影响占比':<10} {'缩放系数':<10} {'0mA 60L/min 修正值'}")
        print("-" * 85)
        for f, _ in file_score.items():
            rate = f"{impact_rate.get(f, 0):.1f}%"
            scale = f"×{scale_suggest.get(f, 1.0):.2f}"
            orig, new = scale_0mA60L.get(f, (0.0, 0.0))
            corr = f"{orig:.2f} → {new:.2f}"
            print(f"{f:<25} {rate:<10} {scale:<10} {corr}")
    else:
        print("✅ 所有产品一致性合格，无需调节")

    print(f"\n🎉 程序执行完毕！")
    print(f"📂 所有结果存放目录：{OUTPUT}")
    input("\n按回车键退出...")