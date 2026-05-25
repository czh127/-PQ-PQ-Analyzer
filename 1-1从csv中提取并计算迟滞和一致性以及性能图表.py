import pandas as pd
import numpy as np
import os
import glob
import matplotlib.pyplot as plt
from openpyxl import Workbook
import warnings

# ============================================================
# 【一、顶级自定义配置区】
# 电流标签位置说明：
# PERF_TEXT_OFFSET_X → 控制标签在曲线终点 右侧多远（正数=右）
# PERF_TEXT_OFFSET_Y → 控制标签 上下波动（正数=上，负数=下，0=居中）
# ============================================================
warnings.filterwarnings('ignore')
CONFIG = {
    # ====================== 基础全局参数 ======================
    "OUTPUT_EXCEL": "产品检测结果总表.xlsx",
    "ENCODING": "gbk",
    "COL_CURRENT": 1,
    "COL_FLOW": 4,
    "COL_PRESS": 10,
    "REF_FLOW": [5.0,10.0,20.0,40.0,60.0],
    "PLOT_FLOW": 20.0,
    "STD_LIMIT": 10.0,
    "STD_LINE1": {"value":10.0,"show":1},
    "TOLERANCE": 1.00,
    "CONSISTENCY_LIMIT_PCT": 5.0,
    "KEY_FLOW": 20.0,
    "HIGH_FLOW": 40.0,
    "KEY_CURRENTS": [0,300,600,900],
    "FILTER_WINDOW": 201,
    "FILTER_ON": True,
    "X_AXIS_EXTEND": 8,
    "HIDE_TOP_RIGHT_BORDER": True,

    # ====================== 1. 单文件性能图表（按你要求） ======================
    "PERF_TITLE_SUFFIX": " 实测",
    "PERF_TITLE_SIZE": 36,
    "PERF_TITLE_BOLD": False,
    "PERF_AXIS_LABEL_SIZE": 28,
    "PERF_AXIS_LABEL_BOLD": False,
    "PERF_CURR_TEXT_SIZE": 28,
    "PERF_CURR_TEXT_BOLD": False,
    "PERF_TICK_LABEL_SIZE": 20,
    "PERF_LINE_WIDTH": 2.0,
    "PERF_TICK_WIDTH": 1.2,
    # 电流标签位置自定义（正右侧 + 上下波动）
    "PERF_TEXT_OFFSET_X": 1.5,     # 水平向右距离（推荐1~3）
    "PERF_TEXT_OFFSET_Y": 0.0,     # 上下波动：0=居中，正数上，负数下

    # ====================== 2. PQ总对比图 ======================
    "PQ_ALL_TITLE_SIZE": 16,
    "PQ_ALL_TITLE_BOLD": True,
    "PQ_ALL_AXIS_SIZE": 12,
    "PQ_ALL_LEGEND_SIZE": 10,
    "PQ_ALL_LEGEND_BOLD": False,
    "PQ_ALL_TICK_LABEL_SIZE": 10,
    "PQ_ALL_LINE_WIDTH": 0.8,
    "PQ_ALL_TICK_WIDTH": 1.2,

    # ====================== 3. 一致性图表 ======================
    "CONSIST_TITLE": "一致性对比",
    "CONSIST_TITLE_SIZE": 16,
    "CONSIST_TITLE_BOLD": True,
    "CONSIST_AXIS_SIZE": 12,
    "CONSIST_AXIS_BOLD": False,
    "CONSIST_TAG_SIZE": 9,
    "CONSIST_TAG_BOLD": False,
    "LABEL_POS_0mA": -18,
    "LABEL_POS_OTHER": 8,
    "CONSIST_TICK_LABEL_SIZE": 10,
    "CONSIST_LINE_WIDTH": 2.0,
    "CONSIST_TICK_WIDTH": 1.2,

    # ====================== 4. 迟滞图表 ======================
    "HYST_TITLE_SIZE": 16,
    "HYST_TITLE_BOLD": True,
    "HYST_AXIS_SIZE": 12,
    "HYST_LEGEND_SIZE": 10,
    "HYST_LEGEND_BOLD": False,
    "HYST_TICK_LABEL_SIZE": 10,
    "HYST_LINE_WIDTH": 1.5,
    "HYST_TICK_WIDTH": 1.2,
};

# ====================== 固定颜色配置 ======================
COLOR_MAP = {0:'#70AD47',1:'#00B0F0',2:'#FFC000',3:'#FF0000',4:'#C00000'}
FILE_COLORS = {
    0:'#00B0F0',1:'#FFC000',2:'#70AD47',3:'#FF0000',4:'#C00000',
    5:'#9933FF',6:'#00FF00',7:'#00FFFF',8:'#FF9900',9:'#999999'
}

# ============================================================
# 【二、工具函数】
# ============================================================
OUTPUT_FOLDER = "检测结果输出"
if not os.path.exists(OUTPUT_FOLDER):
    os.makedirs(OUTPUT_FOLDER)
else:
    count = 1
    while True:
        temp = f"{OUTPUT_FOLDER}_{count}"
        if not os.path.exists(temp):
            OUTPUT_FOLDER = temp
            os.makedirs(OUTPUT_FOLDER)
            break
        count += 1

def save_safe_plot(fig, filename, dpi=150):
    path = os.path.join(OUTPUT_FOLDER, filename)
    base, ext = os.path.splitext(path)
    cnt = 1
    while os.path.exists(path):
        path = f"{base}({cnt}){ext}"
        cnt +=1
    fig.savefig(path, dpi=dpi, bbox_inches='tight')
    return path

def get_closest_index(arr, target):
    return np.argmin(np.abs(np.array(arr)-target))

def safe_percent(numerator, denominator, default=0):
    try:
        return round((numerator/denominator)*100, 1)
    except:
        return default

def moving_average_filter(s, window=201):
    if not CONFIG["FILTER_ON"] or window<=1:
        return s.copy()
    return s.rolling(window=window, center=True, min_periods=1).mean()

def clean_invalid_data(df):
    df = df.replace([None,np.inf,-np.inf], np.nan)
    df = df.dropna(subset=["流量值","压差值"])
    df["流量值"] = pd.to_numeric(df["流量值"], errors="coerce")
    df["压差值"] = pd.to_numeric(df["压差值"], errors="coerce")
    return df

def analyze_consistency_influence_final(df_valid):
    influence = []
    kc = CONFIG["KEY_CURRENTS"]
    kf = CONFIG["KEY_FLOW"]
    hf = CONFIG["HIGH_FLOW"]
    for (i,rf,p),g in df_valid.groupby(["电流值","参考流量值","数据分区"]):
        vals = g["实际压差值"].dropna()
        if len(vals)<2: continue
        mean = vals.mean()
        maxv = vals.max()
        minv = vals.min()
        diff = (maxv-minv)/2
        pct = safe_percent(diff, mean)
        if abs(pct) <= CONFIG["CONSISTENCY_LIMIT_PCT"]: continue
        w = 1.0
        if i in kc:
            if i == 0: bw=10.0
            elif i==300: bw=8.0
            elif i==600: bw=6.0
            elif i==900: bw=4.0
            else: bw=1.0
            w = bw*2 if rf==kf else bw
        elif i>900: w=1.0
        else: w=2.0 if rf>=hf else 1.0
        for f in g["数据源文件"].unique():
            sub = g[g["数据源文件"]==f]
            if sub.empty: continue
            val = sub["实际压差值"].iloc[0]
            dev = abs(val-mean)
            influence.append({"file":f,"I":i,"rf":rf,"P":val,"mean":mean,
                             "weight":w,"wdev":dev*w,"pct":pct})
    if not influence: return None,None,None,None
    df = pd.DataFrame(influence)
    score = df.groupby("file")["wdev"].sum().sort_values(ascending=False)
    total = score.sum()
    rate = {f:round(s/total*100,1) for f,s in score.items()}
    worst = score.index[0]
    scale = {}
    for f in df["file"].unique():
        s = df[df["file"]==f]
        fm = s["P"].mean()
        gm = s["mean"].mean()
        scale[f] = round(gm/fm,3) if fm!=0 else 1.0
    return score, rate, worst, scale

def get_real_0mA60L_press(df_valid, files):
    res = {}
    for f in files:
        sub = df_valid[(df_valid["数据源文件"]==f) &
                       (df_valid["电流值"]==0.0) &
                       (df_valid["参考流量值"]==60.0) &
                       (df_valid["数据分区"]=="前50%")]
        res[f] = round(sub["实际压差值"].iloc[0],2) if not sub.empty else 0.0
    return res

# ============================================================
# 【三、主程序】
# ============================================================
print("="*60)
print("🚀 PQ 综合分析程序")
print("="*60)

os.chdir(os.path.dirname(os.path.abspath(__file__)))
csv_files = glob.glob("*.csv")
print(f"\n📂 扫描到 CSV 文件：{len(csv_files)}")

raw_data = []
pq_compare = []

for file in csv_files:
    try:
        df = pd.read_csv(file, usecols=[1,4,10], encoding=CONFIG["ENCODING"])
        df.columns = ["电流值","流量值","压差值"]
    except: continue
    df = clean_invalid_data(df)
    if df.empty: continue
    df["流量值"] = moving_average_filter(df["流量值"])
    df["压差值"] = moving_average_filter(df["压差值"])
    df = clean_invalid_data(df)
    if df.empty: continue
    fname = os.path.splitext(file)[0]
    for I, g in df.groupby("电流值"):
        I = round(I,2)
        half = g.iloc[:len(g)//2]
        f = half["流量值"].abs().values
        p = half["压差值"].abs().values
        fp, pp = [], []
        for r in CONFIG["REF_FLOW"]:
            idx = get_closest_index(f, r)
            fp.append(round(abs(f[idx]),4))
            pp.append(abs(p[idx]))
        pq_compare.append({"file":fname,"current":I,"flow":fp,"press":pp})
    for I, g in df.groupby("电流值"):
        I = round(I,2)
        mid = len(g)//2
        parts = [("前50%",g.iloc[:mid]),("后50%",g.iloc[mid:])]
        for name, part in parts:
            part = clean_invalid_data(part)
            if part.empty: continue
            f = part["流量值"].abs().values
            p = part["压差值"].abs().values
            for r in CONFIG["REF_FLOW"]:
                idx = get_closest_index(f, r)
                rf = round(abs(f[idx]),4)
                rp = round(abs(p[idx]),4)
                diff = round(abs(rf-r),4)
                ok = "TRUE" if diff<=CONFIG["TOLERANCE"] else "FALSE"
                raw_data.append([file,I,name,r,rp,rf,diff,ok])

df_raw = pd.DataFrame(raw_data, columns=[
    "数据源文件","电流值","数据分区","参考流量值","实际压差值","实际流量值","差值","校验结果"
])
df_valid = df_raw[df_raw["校验结果"]=="TRUE"].copy()

score, rate, worst, scale = analyze_consistency_influence_final(df_valid)
all_files = df_valid["数据源文件"].unique()
orig0 = get_real_0mA60L_press(df_valid, all_files)
scale0 = {f:(orig0.get(f,0.0), round(orig0.get(f,0.0)*scale.get(f,1.0),2)) for f in all_files}

plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

# ============================================================
# 【四、绘图 —— 性能图电流标签固定正右侧】
# ============================================================

# ---------------------- 1. 单文件性能图 ----------------------
for idx, file in enumerate(df_valid["数据源文件"].unique(),1):
    d = df_valid[(df_valid["数据源文件"]==file)&(df_valid["数据分区"]=="前50%")].copy()
    if d.empty: continue
    fig, ax = plt.subplots(figsize=(12,6), dpi=100)
    for curr in sorted(d["电流值"].unique()):
        s = d[d["电流值"]==curr].sort_values("参考流量值")
        ax.plot(s["参考流量值"], s["实际压差值"], marker='o', 
                linewidth=CONFIG["PERF_LINE_WIDTH"], color='#00B0F0')
        # 核心：电流标签固定在【曲线终点正右侧】，可上下左右偏移
        x_last = s["参考流量值"].iloc[-1]
        y_last = s["实际压差值"].iloc[-1]
        ax.text(
            x_last + CONFIG["PERF_TEXT_OFFSET_X"],  # 固定向右
            y_last + CONFIG["PERF_TEXT_OFFSET_Y"],  # 上下波动
            f"{curr}mA",
            fontsize=CONFIG["PERF_CURR_TEXT_SIZE"],
            weight="bold" if CONFIG["PERF_CURR_TEXT_BOLD"] else "normal",
            va="center"  # 垂直居中对齐
        )
    # 样式
    title = f"{os.path.splitext(file)[0]}{CONFIG['PERF_TITLE_SUFFIX']}"
    ax.set_title(title, fontsize=CONFIG["PERF_TITLE_SIZE"],
                 weight="bold" if CONFIG["PERF_TITLE_BOLD"] else "normal")
    ax.set_xlabel("流量 L/min", fontsize=CONFIG["PERF_AXIS_LABEL_SIZE"],
                  weight="bold" if CONFIG["PERF_AXIS_LABEL_BOLD"] else "normal")
    ax.set_ylabel("压差 bar", fontsize=CONFIG["PERF_AXIS_LABEL_SIZE"],
                  weight="bold" if CONFIG["PERF_AXIS_LABEL_BOLD"] else "normal")
    ax.tick_params(axis='both', labelsize=CONFIG["PERF_TICK_LABEL_SIZE"], 
                   width=CONFIG["PERF_TICK_WIDTH"])
    ax.set_xlim(right=70)
    if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
    plt.tight_layout()
    save_safe_plot(fig, f"性能图表_{idx}.png")
    plt.close()

# ---------------------- 2. PQ总对比图 ----------------------
if pq_compare:
    fig, ax = plt.subplots(figsize=(15,8), dpi=100)
    f_list = list({x["file"] for x in pq_compare})
    c_list = sorted({x["current"] for x in pq_compare})
    c_map = {f:FILE_COLORS[i%len(FILE_COLORS)] for i,f in enumerate(f_list)}
    plotted = set()
    for f in f_list:
        dat = [d for d in pq_compare if d["file"]==f]
        for curr in c_list:
            match = next((d for d in dat if d["current"]==curr), None)
            if match:
                lb = f if f not in plotted else ""
                if f not in plotted: plotted.add(f)
                ax.plot(match["flow"], match["press"], marker='o', 
                        linewidth=CONFIG["PQ_ALL_LINE_WIDTH"],
                        color=c_map[f], label=lb, alpha=0.8)
    ax.set_title("所有产品PQ曲线对比", fontsize=CONFIG["PQ_ALL_TITLE_SIZE"],
                 weight="bold" if CONFIG["PQ_ALL_TITLE_BOLD"] else "normal")
    ax.set_xlabel("流量 L/min", fontsize=CONFIG["PQ_ALL_AXIS_SIZE"])
    ax.set_ylabel("压差 bar", fontsize=CONFIG["PQ_ALL_AXIS_SIZE"])
    ax.grid(alpha=0.3)
    ax.set_xlim(0, max(CONFIG["REF_FLOW"])+10)
    ax.tick_params(axis='both', labelsize=CONFIG["PQ_ALL_TICK_LABEL_SIZE"],
                   width=CONFIG["PQ_ALL_TICK_WIDTH"])
    if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
    leg = ax.legend(bbox_to_anchor=(1.05,1), loc="upper left")
    for t in leg.get_texts():
        t.set_fontsize(CONFIG["PQ_ALL_LEGEND_SIZE"])
        t.set_weight("bold" if CONFIG["PQ_ALL_LEGEND_BOLD"] else "normal")
    plt.tight_layout()
    save_safe_plot(fig, "所有产品PQ对比图.png")
    plt.close()

# ---------------------- 3. 迟滞图 ----------------------
hyst_data = []
for (f,i,rf),g in df_valid.groupby(["数据源文件","电流值","参考流量值"]):
    b = g[g["数据分区"]=="前50%"]["实际压差值"]
    a = g[g["数据分区"]=="后50%"]["实际压差值"]
    if len(b)==1 and len(a)==1:
        hyst_data.append([f,i,rf,round(abs(b.iloc[0]-a.iloc[0]),2)])
df_hyst = pd.DataFrame(hyst_data, columns=["文件","电流","流量","迟滞"])

fig, ax = plt.subplots(figsize=(15,6), dpi=100)
hp = df_hyst[df_hyst["流量"]==CONFIG["PLOT_FLOW"]].sort_values(["文件","电流"])
for f in hp["文件"].unique():
    s = hp[hp["文件"]==f]
    ax.plot(s["电流"], s["迟滞"], marker='o', label=f, linewidth=CONFIG["HYST_LINE_WIDTH"])
if CONFIG["STD_LINE1"]["show"]:
    val = CONFIG["STD_LINE1"]["value"]
    ax.plot(hp["电流"].unique(), [val]*len(hp["电流"].unique()), 'r--', 
            linewidth=CONFIG["HYST_LINE_WIDTH"], label=f"标准线{val}")
ax.set_title("迟滞特性", fontsize=CONFIG["HYST_TITLE_SIZE"],
             weight="bold" if CONFIG["HYST_TITLE_BOLD"] else "normal")
ax.set_xlabel("电流", fontsize=CONFIG["HYST_AXIS_SIZE"])
ax.set_ylabel("迟滞 bar", fontsize=CONFIG["HYST_AXIS_SIZE"])
ax.grid(alpha=0.3)
ax.tick_params(axis='both', labelsize=CONFIG["HYST_TICK_LABEL_SIZE"],
               width=CONFIG["HYST_TICK_WIDTH"])
lgh = ax.legend(bbox_to_anchor=(1,1))
for t in lgh.get_texts():
    t.set_fontsize(CONFIG["HYST_LEGEND_SIZE"])
    t.set_weight("bold" if CONFIG["HYST_LEGEND_BOLD"] else "normal")
if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
plt.tight_layout()
save_safe_plot(fig, "迟滞图.png")
plt.close()

# ---------------------- 4. 一致性图 ----------------------
cons_data = []
for (i,rf,p),g in df_valid.groupby(["电流值","参考流量值","数据分区"]):
    v = g["实际压差值"].dropna()
    if len(v)<1: continue
    maxv = v.max()
    minv = v.min()
    avg = (maxv+minv)/2
    diff = (maxv-minv)/2
    cons_data.append([round(i,2), rf, p, maxv, minv, avg, diff, safe_percent(diff,avg)])
df_cons = pd.DataFrame(cons_data, columns=["电流","流量","分区","max","min","avg","diff","pct"])

fig, ax = plt.subplots(figsize=(15,6), dpi=100)
for curr in sorted(df_cons["电流"].unique()):
    s = df_cons[(df_cons["电流"]==curr)&(df_cons["分区"]=="前50%")].sort_values("流量")
    ax.plot(s["流量"], s["avg"], marker='o', color='#00B0F0', 
            linewidth=CONFIG["CONSIST_LINE_WIDTH"])
for curr in sorted(df_cons["电流"].unique()):
    s = df_cons[(df_cons["电流"]==curr)&(df_cons["分区"]=="前50%")].sort_values("流量")
    for x,y,d,p in zip(s["流量"],s["avg"],s["diff"],s["pct"]):
        lab = f"±{d}bar" if curr==0 and x in [5,10] else f"±{p}%"
        oy = CONFIG["LABEL_POS_0mA"] if curr==0 else CONFIG["LABEL_POS_OTHER"]
        ax.annotate(lab, (x,y), xytext=(0,oy), textcoords='offset points', ha='center',
                    fontsize=CONFIG["CONSIST_TAG_SIZE"],
                    weight="bold" if CONFIG["CONSIST_TAG_BOLD"] else "normal")
ax.set_title(CONFIG["CONSIST_TITLE"], fontsize=CONFIG["CONSIST_TITLE_SIZE"],
             weight="bold" if CONFIG["CONSIST_TITLE_BOLD"] else "normal")
ax.set_xlabel("流量 L/min", fontsize=CONFIG["CONSIST_AXIS_SIZE"],
              weight="bold" if CONFIG["CONSIST_AXIS_BOLD"] else "normal")
ax.set_ylabel("压差 bar", fontsize=CONFIG["CONSIST_AXIS_SIZE"],
              weight="bold" if CONFIG["CONSIST_AXIS_BOLD"] else "normal")
ax.set_xlim(right=max(CONFIG["REF_FLOW"])+CONFIG["X_AXIS_EXTEND"])
ax.tick_params(axis='both', labelsize=CONFIG["CONSIST_TICK_LABEL_SIZE"],
               width=CONFIG["CONSIST_TICK_WIDTH"])
if CONFIG["HIDE_TOP_RIGHT_BORDER"]:
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
plt.tight_layout()
save_safe_plot(fig, "一致性图.png")
plt.close()

# ============================================================
# 【五、Excel导出】
# ============================================================
wb = Workbook()
wb.remove(wb.active)
ws1 = wb.create_sheet("原始数据")
ws1.append(["文件","电流","分区","参考流量","压差","流量","误差","校验"])
for r in raw_data: ws1.append(r)
ws2 = wb.create_sheet("迟滞数据")
ws2.append(["文件","电流","流量","迟滞"])
for r in hyst_data: ws2.append(r)
ws3 = wb.create_sheet("一致性数据")
ws3.append(["电流","流量","分区","最大值","最小值","平均值","半差值","一致性%"])
for r in cons_data: ws3.append(r)
wb.save(os.path.join(OUTPUT_FOLDER, CONFIG["OUTPUT_EXCEL"]))

print("\n✅ 分析完成！结果保存在：" + OUTPUT_FOLDER)
print("="*60)